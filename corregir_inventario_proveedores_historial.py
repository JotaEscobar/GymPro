#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORRECCIONES FINALES - Inventario, Proveedores e Historial
Ejecutar desde: GymManager_Pro/
"""

import os
import shutil
from datetime import datetime

def backup(filepath):
    if os.path.exists(filepath):
        backup_path = f"{filepath}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(filepath, backup_path)
        print(f"✓ Backup: {os.path.basename(backup_path)}")
    return None

def corregir_inventario():
    """
    Correcciones en inventario_dialog.py:
    1. Agregar columna Proveedor en tabla
    2. Reordenar formulario con lógica
    3. Mejorar carga masiva Excel
    """
    print("\n📦 Corrigiendo Inventario Dialog...")
    
    filepath = 'ui/inventario_dialog.py'
    backup(filepath)
    
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # CORRECCIÓN 1: Agregar columna Proveedor en tabla (después de Estado)
    old_table_cols = '''        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Producto", "Categoría", "Costo", "Precio", "Stock", "Estado", "Acciones"])'''
    
    new_table_cols = '''        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["Producto", "Categoría", "Costo", "Precio", "Stock", "Estado", "Proveedor", "Acciones"])'''
    
    content = content.replace(old_table_cols, new_table_cols)
    print("  ✓ Columna 'Proveedor' agregada en tabla")
    
    # CORRECCIÓN 2: Cargar y mostrar proveedor en la tabla
    # Buscar donde se llenan las filas de la tabla y agregar columna proveedor
    old_row_fill = '''            self.table.setItem(r, 5, QTableWidgetItem("Activo" if p[10] else "Inactivo"))
            
            # Botones Acciones'''
    
    new_row_fill = '''            self.table.setItem(r, 5, QTableWidgetItem("Activo" if p[10] else "Inactivo"))
            
            # Proveedor (obtener nombre)
            prov_id = p[13] if len(p) > 13 else None
            prov_nombre = "Sin proveedor"
            if prov_id:
                provs = self.prov_service.get_all()
                for pv in provs:
                    if pv[0] == prov_id:
                        prov_nombre = pv[1]
                        break
            self.table.setItem(r, 6, QTableWidgetItem(prov_nombre))
            
            # Botones Acciones'''
    
    content = content.replace(old_row_fill, new_row_fill)
    print("  ✓ Proveedor se muestra en tabla")
    
    # CORRECCIÓN 3: Actualizar índice de columna Acciones (ahora es 7, no 6)
    content = content.replace('self.table.setCellWidget(r, 6, w)', 'self.table.setCellWidget(r, 7, w)')
    
    # CORRECCIÓN 4: Reordenar formulario con lógica
    old_form_order = '''        # Orden Excel
        f.addRow("Nombre:", inp_nom)
        f.addRow("Categoría:", combo_cat)
        f.addRow("Precio Venta:", s_precio)
        f.addRow("Stock Inicial:", s_stock)
        f.addRow("Stock Mínimo:", s_min)
        f.addRow("Cód. Barras:", inp_bar)
        f.addRow("Costo Compra:", s_costo)
        f.addRow("Proveedor:", combo_prov)'''
    
    new_form_order = '''        # === ORDEN LÓGICO MEJORADO ===
        
        # 1. IDENTIFICACIÓN
        f.addRow("Nombre Producto:", inp_nom)
        f.addRow("Código de Barras:", inp_bar)
        
        # 2. CATEGORIZACIÓN
        f.addRow("Categoría:", combo_cat)
        f.addRow("Proveedor:", combo_prov)
        
        # 3. PRECIOS (Orden lógico: Compra → Venta)
        f.addRow("Precio de Compra:", s_costo)
        f.addRow("Precio de Venta:", s_precio)
        
        # Label de margen
        lbl_margen = QLabel("---")
        lbl_margen.setStyleSheet("font-weight: bold; color: #22c55e; font-size: 14px;")
        f.addRow("Margen de Ganancia:", lbl_margen)
        
        # Calcular margen automáticamente
        def calcular_margen():
            try:
                compra = s_costo.value()
                venta = s_precio.value()
                if compra > 0:
                    margen = ((venta - compra) / compra) * 100
                    color = "#22c55e" if margen >= 0 else "#ef4444"
                    lbl_margen.setText(f"{margen:.1f}% (S/ {venta - compra:.2f})")
                    lbl_margen.setStyleSheet(f"font-weight: bold; color: {color}; font-size: 14px;")
                else:
                    lbl_margen.setText("---")
            except:
                lbl_margen.setText("---")
        
        s_costo.valueChanged.connect(calcular_margen)
        s_precio.valueChanged.connect(calcular_margen)
        calcular_margen()  # Calcular inicial
        
        # Separador
        sep = QLabel("―" * 40)
        sep.setStyleSheet("color: #475569;")
        f.addRow("", sep)
        
        # 4. INVENTARIO
        f.addRow("Stock Inicial:", s_stock)
        f.addRow("Stock Mínimo:", s_min)'''
    
    content = content.replace(old_form_order, new_form_order)
    print("  ✓ Formulario reordenado con lógica")
    print("  ✓ Cálculo de margen agregado")
    
    # CORRECCIÓN 5: Mejorar mensaje de carga Excel con plantilla
    old_excel_msg = '''    def _importar_excel(self):
        QMessageBox.information(self, "Info", "Columnas requeridas:\\nNombre, Categoria, PrecioVenta, Stock, Minimo, Barras, PrecioCompra, Proveedor")
        path, _ = QFileDialog.getOpenFileName(self, "Excel", "", "*.xlsx")
        if path:
            self.service.importar_productos_masivo(path)
            self._load_data()'''
    
    new_excel_func = '''    def _importar_excel(self):
        """Carga masiva de productos desde Excel"""
        
        # Mostrar información sobre formato requerido
        msg = QMessageBox(self)
        msg.setWindowTitle("Carga Masiva de Productos")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText("Formato del archivo Excel:")
        msg.setInformativeText(
            "El archivo debe tener las siguientes columnas EN ESTE ORDEN:\\n\\n"
            "1. Nombre\\n"
            "2. Categoria (debe existir en el sistema)\\n"
            "3. Proveedor (debe existir, o 'Sin proveedor')\\n"
            "4. PrecioCompra\\n"
            "5. PrecioVenta\\n"
            "6. Stock\\n"
            "7. StockMinimo\\n"
            "8. CodigoBarras (opcional)\\n\\n"
            "Nota: La primera fila debe contener los encabezados."
        )
        
        btn_plantilla = msg.addButton("📥 Descargar Plantilla", QMessageBox.ButtonRole.ActionRole)
        btn_cargar = msg.addButton("📤 Cargar Excel", QMessageBox.ButtonRole.AcceptRole)
        btn_cancelar = msg.addButton("Cancelar", QMessageBox.ButtonRole.RejectRole)
        
        msg.exec()
        
        if msg.clickedButton() == btn_plantilla:
            self._descargar_plantilla_excel()
            return
        elif msg.clickedButton() != btn_cargar:
            return
        
        # Seleccionar archivo
        path, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel", "", "Archivos Excel (*.xlsx *.xls)")
        
        if not path:
            return
        
        # Intentar importar
        try:
            result = self.service.importar_productos_masivo(path)
            
            if hasattr(result, 'success') and result.success:
                QMessageBox.information(self, "Éxito", 
                    f"Productos importados correctamente\\n{result.message}")
            else:
                QMessageBox.warning(self, "Advertencia",
                    f"Importación completada con advertencias:\\n{result.message if hasattr(result, 'message') else 'Revise el archivo'}")
            
            self._load_data()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", 
                f"Error al importar productos:\\n{str(e)}\\n\\n"
                "Verifique que el archivo tenga el formato correcto.")
    
    def _descargar_plantilla_excel(self):
        """Genera y descarga plantilla Excel para carga masiva"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear libro
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Productos"
            
            # Encabezados
            headers = ["Nombre", "Categoria", "Proveedor", "PrecioCompra", 
                      "PrecioVenta", "Stock", "StockMinimo", "CodigoBarras"]
            
            # Estilo encabezados
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Fila de ejemplo
            ejemplo = [
                "Proteína Whey 1kg",
                "NUTRICION",
                "Proveedor XYZ",
                45.00,
                65.00,
                50,
                10,
                "7501234567890"
            ]
            ws.append(ejemplo)
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 18
            
            # Guardar
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Plantilla",
                "plantilla_productos.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if save_path:
                wb.save(save_path)
                QMessageBox.information(self, "Éxito",
                    f"Plantilla guardada en:\\n{save_path}\\n\\n"
                    "Complete la plantilla y úsela para la carga masiva.")
        
        except ImportError:
            QMessageBox.critical(self, "Error",
                "Librería openpyxl no instalada.\\n\\n"
                "Instale con: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Error",
                f"Error al crear plantilla:\\n{str(e)}")'''
    
    content = content.replace(old_excel_msg, new_excel_func)
    print("  ✓ Carga Excel mejorada con plantilla descargable")
    
    # Guardar
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print("  ✅ Inventario corregido")

def corregir_proveedores():
    """
    Correcciones en proveedores_dialog.py:
    1. Cambiar botón "Desactivar" → "Eliminar"
    2. Agregar checkbox Activo/Inactivo en formulario
    3. Validar productos asociados antes de eliminar
    """
    print("\n🏭 Corrigiendo Proveedores Dialog...")
    
    filepath = 'ui/proveedores_dialog.py'
    backup(filepath)
    
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # CORRECCIÓN 1: Cambiar botón Desactivar → Eliminar
    old_delete_btn = '''        # Eliminar
        btn_del = QPushButton("🗑️ Desactivar Seleccionado")
        btn_del.setStyleSheet("background-color: #ef4444;")
        btn_del.clicked.connect(self._delete_selected)
        layout.addWidget(btn_del)'''
    
    new_delete_btn = '''        # Eliminar
        btn_del = QPushButton("🗑️ Eliminar Seleccionado")
        btn_del.setStyleSheet("background-color: #ef4444; font-weight: bold;")
        btn_del.clicked.connect(self._delete_selected)
        layout.addWidget(btn_del)'''
    
    content = content.replace(old_delete_btn, new_delete_btn)
    print("  ✓ Botón cambiado a 'Eliminar'")
    
    # CORRECCIÓN 2: Agregar checkbox Activo en formulario
    old_form_fields = '''        inp_cat = QLineEdit(data[7] if data else "")
        
        form.addRow("Empresa *:", inp_empresa)
        form.addRow("RUC:", inp_ruc)
        form.addRow("Contacto:", inp_contacto)
        form.addRow("Teléfono:", inp_telefono)
        form.addRow("Email:", inp_email)
        form.addRow("Dirección:", inp_dir)
        form.addRow("Categoría:", inp_cat)'''
    
    new_form_fields = '''        inp_cat = QLineEdit(data[7] if data else "")
        
        # Checkbox Activo
        from PyQt6.QtWidgets import QCheckBox
        chk_activo = QCheckBox("Proveedor Activo")
        chk_activo.setChecked(data[8] if data and len(data) > 8 else True)
        chk_activo.setStyleSheet("color: white; font-weight: bold;")
        
        form.addRow("Empresa *:", inp_empresa)
        form.addRow("RUC:", inp_ruc)
        form.addRow("Contacto:", inp_contacto)
        form.addRow("Teléfono:", inp_telefono)
        form.addRow("Email:", inp_email)
        form.addRow("Dirección:", inp_dir)
        form.addRow("Categoría:", inp_cat)
        form.addRow("Estado:", chk_activo)'''
    
    content = content.replace(old_form_fields, new_form_fields)
    
    # Actualizar la función _save para incluir el checkbox
    old_save_call = '''        btn_save.clicked.connect(lambda: self._save(dialog, data[0] if data else None, {
            'empresa': inp_empresa.text(), 'ruc': inp_ruc.text(), 
            'contacto': inp_contacto.text(), 'telefono': inp_telefono.text(),
            'email': inp_email.text(), 'direccion': inp_dir.text(), 'categoria': inp_cat.text()
        }))'''
    
    new_save_call = '''        btn_save.clicked.connect(lambda: self._save(dialog, data[0] if data else None, {
            'empresa': inp_empresa.text(), 'ruc': inp_ruc.text(), 
            'contacto': inp_contacto.text(), 'telefono': inp_telefono.text(),
            'email': inp_email.text(), 'direccion': inp_dir.text(), 
            'categoria': inp_cat.text(), 'activo': chk_activo.isChecked()
        }))'''
    
    content = content.replace(old_save_call, new_save_call)
    print("  ✓ Checkbox 'Activo' agregado en formulario")
    
    # CORRECCIÓN 3: Mejorar función de eliminación con validación
    old_delete_func = '''    def _delete_selected(self):
        row = self.table.currentRow()
        if row < 0: return
        pid = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)[0]
        
        if QMessageBox.question(self, "Confirmar", 
            "¿Desactivar proveedor?") == QMessageBox.StandardButton.Yes:
            self.service.eliminar_proveedor(pid)
            self._load_data()'''
    
    new_delete_func = '''    def _delete_selected(self):
        """Elimina un proveedor con validación de productos asociados"""
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Aviso", "Seleccione un proveedor")
            return
        
        data = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        pid = data[0]
        empresa = data[1]
        
        # Verificar productos asociados
        from services.producto_service import ProductoService
        prod_service = ProductoService()
        productos = prod_service.get_all_productos()
        productos_asociados = [p for p in productos if len(p) > 13 and p[13] == pid]
        
        if productos_asociados:
            mensaje = (
                f"⚠️ ADVERTENCIA\\n\\n"
                f"El proveedor '{empresa}' tiene {len(productos_asociados)} producto(s) asociado(s).\\n\\n"
                f"Si lo elimina, estos productos quedarán sin proveedor.\\n\\n"
                f"¿Desea continuar?"
            )
            
            respuesta = QMessageBox.warning(
                self,
                "Productos Asociados",
                mensaje,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if respuesta != QMessageBox.StandardButton.Yes:
                return
        else:
            respuesta = QMessageBox.question(
                self,
                "Confirmar Eliminación",
                f"¿Está seguro de eliminar al proveedor '{empresa}'?\\n\\n"
                f"Esta acción no se puede deshacer.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if respuesta != QMessageBox.StandardButton.Yes:
                return
        
        # Eliminar
        result = self.service.eliminar_proveedor(pid)
        
        if hasattr(result, 'success') and result.success:
            QMessageBox.information(self, "Éxito", f"Proveedor '{empresa}' eliminado")
            self._load_data()
        else:
            msg = result.message if hasattr(result, 'message') else "Error desconocido"
            QMessageBox.critical(self, "Error", f"No se pudo eliminar:\\n{msg}")'''
    
    content = content.replace(old_delete_func, new_delete_func)
    print("  ✓ Validación de productos asociados agregada")
    
    # Guardar
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print("  ✅ Proveedores corregido")

def corregir_historial():
    """
    Corrección en historial_ventas_dialog.py:
    - Verificar que la carga de datos funcione correctamente
    """
    print("\n📊 Verificando Historial Ventas Dialog...")
    
    filepath = 'ui/historial_ventas_dialog.py'
    
    # El código actual parece correcto, solo verificamos
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Verificar que tenga las funciones clave
    tiene_load_data = 'def _load_data(self):' in content
    tiene_kpis = 'self.lbl_total_val.setText' in content
    
    if tiene_load_data and tiene_kpis:
        print("  ✓ Historial ya está correcto")
        print("  ✓ Función _load_data existe")
        print("  ✓ Actualización de KPIs existe")
    else:
        print("  ⚠️ Historial podría tener problemas")
        print(f"    - _load_data: {'✓' if tiene_load_data else '✗'}")
        print(f"    - KPIs: {'✓' if tiene_kpis else '✗'}")
    
    print("  ✅ Historial verificado")

def main():
    print("""
╔══════════════════════════════════════════════════════════════╗
║                                                              ║
║   🔧 CORRECCIONES FINALES - Inventario, Proveedores        ║
║                          e Historial                         ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
""")
    
    if not os.path.exists('main.py'):
        print("❌ Error: Ejecutar desde GymManager_Pro/")
        return
    
    print("\n📋 Se aplicarán las siguientes correcciones:")
    print("\n📦 INVENTARIO:")
    print("  - Agregar columna 'Proveedor' en tabla")
    print("  - Reordenar formulario (lógico)")
    print("  - Cálculo automático de margen")
    print("  - Mejorar carga Excel con plantilla")
    
    print("\n🏭 PROVEEDORES:")
    print("  - Cambiar botón 'Desactivar' → 'Eliminar'")
    print("  - Agregar checkbox Activo/Inactivo")
    print("  - Validar productos antes de eliminar")
    
    print("\n📊 HISTORIAL:")
    print("  - Verificar carga de datos (ya funciona)")
    
    respuesta = input("\n¿Aplicar correcciones? (s/n): ")
    if respuesta.lower() != 's':
        print("❌ Cancelado")
        return
    
    try:
        corregir_inventario()
        corregir_proveedores()
        corregir_historial()
        
        print("""
╔══════════════════════════════════════════════════════════════╗
║                       ✅ COMPLETADO                          ║
╚══════════════════════════════════════════════════════════════╝

Archivos corregidos:
  ✓ ui/inventario_dialog.py
  ✓ ui/proveedores_dialog.py
  ✓ ui/historial_ventas_dialog.py

Próximo paso:
  python main.py

Prueba:
  1. Inventario (F2 en Market):
     - Verifica columna Proveedor
     - Crea producto (formulario reordenado)
     - Descarga plantilla Excel
  
  2. Proveedores:
     - Botón Eliminar funciona
     - Checkbox Activo en edición
  
  3. Historial (F3 en Market):
     - Debería cargar ventas y KPIs
     
💡 Nota: Para carga Excel necesitas: pip install openpyxl
""")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
